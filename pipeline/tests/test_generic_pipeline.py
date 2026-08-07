from io import BytesIO
from datetime import datetime
from pathlib import Path
import shutil

from fastapi.testclient import TestClient
from openpyxl import Workbook
import pytest
from rdflib import Graph, URIRef

from automating_alignments import automated_alignments
from pipeline import generic_pipeline
from pipeline import playground_server
from pipeline.run_store import RunStore
from triple_store_ingestion import ingest as fuseki_ingest


def test_upload_csv_creates_run_and_preview(tmp_path, monkeypatch):
    monkeypatch.setattr(playground_server, "RUN_STORE", RunStore(tmp_path / "runs"))
    client = TestClient(playground_server.app)

    response = client.post(
        "/api/runs",
        files={"file": ("inventory.csv", b"sku,price\nA-1,10\nB-2,20\n", "text/csv")},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["stages"]["upload"]["status"] == "success"
    assert payload["source"]["stored_filename"] == "inventory.csv"
    assert payload["source"]["preview"]["columns"] == ["sku", "price"]
    assert payload["source"]["preview"]["total_rows"] == 2
    assert "relative_path" not in payload["artifacts"][0]


def test_upload_xlsx_creates_preview_and_rml_csv_source(tmp_path, monkeypatch):
    run_store = RunStore(tmp_path / "runs")
    monkeypatch.setattr(playground_server, "RUN_STORE", run_store)
    client = TestClient(playground_server.app)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Inventory"
    worksheet.append(["sku", "price", "available"])
    worksheet.append(["A-1", 10.5, True])
    worksheet.append(["B-2", 20, False])
    workbook.create_sheet("Notes").append(["This sheet is not active"])
    content = BytesIO()
    workbook.save(content)
    workbook.close()

    response = client.post(
        "/api/runs",
        files={
            "file": (
                "inventory.xlsx",
                content.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["stages"]["upload"]["status"] == "success"
    assert payload["source"]["stored_filename"] == "inventory.xlsx"
    assert payload["source"]["mapping_source_filename"] == "inventory.csv"
    assert payload["source"]["preview"]["format"] == "xlsx"
    assert payload["source"]["preview"]["sheet_name"] == "Inventory"
    assert payload["source"]["preview"]["columns"] == ["sku", "price", "available"]
    assert payload["source"]["preview"]["rows"][0] == {
        "sku": "A-1",
        "price": "10.5",
        "available": "True",
    }
    assert {artifact["kind"] for artifact in payload["artifacts"]} == {"xlsx", "csv"}

    saved = run_store.load(payload["id"])
    mapping_source = run_store.resolve_relative(payload["id"], saved["source"]["relative_path"])
    assert mapping_source.name == "inventory.csv"
    assert mapping_source.read_text(encoding="utf-8").splitlines() == [
        "sku,price,available",
        "A-1,10.5,True",
        "B-2,20,False",
    ]

    workbook_artifact = next(item for item in payload["artifacts"] if item["kind"] == "xlsx")
    preview_response = client.get(workbook_artifact["preview_url"])
    assert preview_response.status_code == 200
    assert preview_response.json()["table"]["sheet_name"] == "Inventory"


def test_upload_xlsx_normalizes_multirow_headers_and_date_time(tmp_path, monkeypatch):
    run_store = RunStore(tmp_path / "runs")
    monkeypatch.setattr(playground_server, "RUN_STORE", run_store)
    client = TestClient(playground_server.app)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Measurements"
    worksheet.append([None, None, "Ambient temperature"])
    worksheet.append(["Date", "Time", "SENSOR_A_TEMPERATURE"])
    worksheet.append(["unit", None, "degC"])
    worksheet.append([datetime(2025, 1, 1), "00:00:00", 4.679602775605543])
    worksheet.append([datetime(2025, 1, 1), "00:15:00", 4.023384243091742])
    content = BytesIO()
    workbook.save(content)
    workbook.close()

    response = client.post(
        "/api/runs",
        files={
            "file": (
                "data.xlsx",
                content.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    payload = response.json()
    value_column = "Ambient temperature | SENSOR_A_TEMPERATURE | degC"
    assert payload["source"]["preview"]["columns"] == ["DateTime", value_column]
    assert payload["source"]["preview"]["header_row_count"] == 3
    assert payload["source"]["preview"]["data_start_row"] == 4
    assert payload["source"]["preview"]["rows"] == [
        {"DateTime": "2025-01-01T00:00:00", value_column: "4.679602775605543"},
        {"DateTime": "2025-01-01T00:15:00", value_column: "4.023384243091742"},
    ]

    saved = run_store.load(payload["id"])
    mapping_source = run_store.resolve_relative(payload["id"], saved["source"]["relative_path"])
    assert mapping_source.read_text(encoding="utf-8").splitlines() == [
        f"DateTime,{value_column}",
        "2025-01-01T00:00:00,4.679602775605543",
        "2025-01-01T00:15:00,4.023384243091742",
    ]


def test_upload_rejects_unsupported_tabular_extension(tmp_path, monkeypatch):
    monkeypatch.setattr(playground_server, "RUN_STORE", RunStore(tmp_path / "runs"))
    client = TestClient(playground_server.app)

    response = client.post(
        "/api/runs",
        files={"file": ("inventory.xls", b"legacy workbook", "application/vnd.ms-excel")},
    )

    assert response.status_code == 400
    assert response.json()["detail"] == "Please upload a .csv or .xlsx file."


def test_delete_run_removes_uploaded_file_and_artifacts(tmp_path, monkeypatch):
    run_store = RunStore(tmp_path / "runs")
    monkeypatch.setattr(playground_server, "RUN_STORE", run_store)
    client = TestClient(playground_server.app)
    upload = client.post(
        "/api/runs",
        files={"file": ("inventory.csv", b"sku,price\nA-1,10\n", "text/csv")},
    )
    run_id = upload.json()["id"]
    run_directory = run_store.run_dir(run_id)
    assert run_directory.is_dir()

    response = client.delete(f"/api/runs/{run_id}")

    assert response.status_code == 204
    assert response.content == b""
    assert not run_directory.exists()
    assert client.get(f"/api/runs/{run_id}").status_code == 404


def test_graph_name_is_converted_to_a_safe_named_graph_uri():
    uri = generic_pipeline.graph_name_to_uri("Quarterly products")
    assert uri == "https://example.org/graphs/Quarterly%20products"
    assert generic_pipeline.graph_name_to_uri("https://data.example/graphs/items") == (
        "https://data.example/graphs/items"
    )


def test_ingest_graph_clears_named_graph_before_upload(tmp_path, monkeypatch):
    rdf_path = tmp_path / "mapped.ttl"
    rdf_path.write_text("<https://example.org/s> <https://example.org/p> <https://example.org/o> .")
    calls = []

    monkeypatch.setattr(
        generic_pipeline.ingest,
        "delete_graph",
        lambda graph_uri: calls.append(("clear", graph_uri)) or True,
    )
    monkeypatch.setattr(
        generic_pipeline.ingest,
        "upload_graph",
        lambda path, graph_uri: calls.append(("upload", graph_uri, path)) or True,
    )

    result = generic_pipeline.ingest_graph(rdf_path, "replacement graph")

    graph_uri = "https://example.org/graphs/replacement%20graph"
    assert calls == [("clear", graph_uri), ("upload", graph_uri, str(rdf_path))]
    assert result == {"graph_uri": graph_uri, "graph_cleared": True}


def test_unit_alignment_converts_a_copy_and_preserves_mapped_rdf(tmp_path):
    mapped_path = tmp_path / "mapped.ttl"
    mapped_path.write_text(
        """
        @prefix qudt: <http://qudt.org/schema/qudt/> .
        @prefix sosa: <http://www.w3.org/ns/sosa/> .
        @prefix unit: <http://qudt.org/vocab/unit/> .
        @prefix xsd: <http://www.w3.org/2001/XMLSchema#> .

        <urn:observation:1> a sosa:Observation ;
            sosa:hasSimpleResult "1000"^^xsd:double ;
            qudt:hasUnit unit:MicroS-PER-CentiM .
        <urn:observation:2> a sosa:Observation ;
            sosa:hasSimpleResult "2500"^^xsd:double ;
            qudt:hasUnit unit:MicroS-PER-CentiM .
        """,
        encoding="utf-8",
    )
    original = mapped_path.read_text(encoding="utf-8")

    result = generic_pipeline.run_unit_alignment(
        tmp_path,
        mapped_path,
        "https://qudt.org/vocab/unit/MilliS-PER-CentiM",
    )

    assert mapped_path.read_text(encoding="utf-8") == original
    assert result["converted_observations"] == 2
    assert result["aligned_observations"] == 2
    assert result["target_unit"] == "http://qudt.org/vocab/unit/MilliS-PER-CentiM"
    aligned = Graph().parse(result["output_path"], format="turtle")
    target = URIRef(result["target_unit"])
    subjects = set(aligned.subjects(predicate=automated_alignments.SOSA.hasSimpleResult))
    assert sorted(
        float(aligned.value(subject, automated_alignments.SOSA.hasSimpleResult))
        for subject in subjects
    ) == [1.0, 2.5]
    assert all(aligned.value(subject, automated_alignments.QUDT.hasUnit) == target for subject in subjects)


@pytest.mark.parametrize(
    ("source_name", "target_name", "value", "expected"),
    [
        ("MicroS-PER-M", "MilliS-PER-M", 1000.0, 1.0),
        ("MilliGM-PER-L", "GM-PER-L", 1000.0, 1.0),
        ("PPTH", "PPM", 1.0, 1000.0),
        ("DEG_C", "K", 20.0, 293.15),
        ("K", "DEG_C", 293.15, 20.0),
    ],
)
def test_known_water_measurement_conversions(source_name, target_name, value, expected):
    unit_root = "http://qudt.org/vocab/unit/"
    source = automated_alignments.KNOWN_CONVERSIONS[URIRef(unit_root + source_name)]
    target = automated_alignments.KNOWN_CONVERSIONS[URIRef(unit_root + target_name)]

    converted = automated_alignments.convert_qudt_value(value, *source, *target)

    assert converted == pytest.approx(expected)


def test_unit_alignment_rejects_incompatible_known_families(tmp_path):
    mapped_path = tmp_path / "mapped.ttl"
    mapped_path.write_text(
        """
        @prefix qudt: <http://qudt.org/schema/qudt/> .
        @prefix sosa: <http://www.w3.org/ns/sosa/> .
        @prefix unit: <http://qudt.org/vocab/unit/> .
        <urn:observation> sosa:hasSimpleResult 1000 ; qudt:hasUnit unit:MicroS-PER-CentiM .
        """,
        encoding="utf-8",
    )

    with pytest.raises(generic_pipeline.PipelineError, match="electrical conductivity"):
        generic_pipeline.run_unit_alignment(
            tmp_path,
            mapped_path,
            "http://qudt.org/vocab/unit/MilliGM-PER-L",
        )


def test_ingest_graph_aborts_when_named_graph_cannot_be_cleared(tmp_path, monkeypatch):
    rdf_path = tmp_path / "mapped.ttl"
    rdf_path.write_text("<https://example.org/s> <https://example.org/p> <https://example.org/o> .")
    upload_called = False

    monkeypatch.setattr(generic_pipeline.ingest, "delete_graph", lambda _graph_uri: False)

    def upload_graph(_path, _graph_uri):
        nonlocal upload_called
        upload_called = True
        return True

    monkeypatch.setattr(generic_pipeline.ingest, "upload_graph", upload_graph)

    with pytest.raises(generic_pipeline.PipelineError, match="No data was uploaded"):
        generic_pipeline.ingest_graph(rdf_path, "replacement graph")
    assert upload_called is False


def test_fuseki_query_leaves_named_graph_selection_to_query(monkeypatch):
    captured = {}

    class Response:
        ok = True
        status_code = 200
        text = ""
        headers = {"Content-Type": "application/sparql-results+json; charset=utf-8"}

        @staticmethod
        def json():
            return {
                "head": {"vars": ["subject"]},
                "results": {
                    "bindings": [
                        {"subject": {"type": "uri", "value": "https://example.org/item"}}
                    ]
                },
            }

    def post(url, **kwargs):
        captured.update({"url": url, **kwargs})
        return Response()

    monkeypatch.setattr(fuseki_ingest.requests, "post", post)

    query = """
        SELECT ?subject
        WHERE {
          GRAPH <https://example.org/graphs/items> {
            ?subject ?predicate ?object
          }
        }
    """
    result = fuseki_ingest.query_graph(query)

    assert captured["data"] == {"query": query}
    assert "default-graph-uri" not in captured["data"]
    assert result["type"] == "select"
    assert result["variables"] == ["subject"]
    assert result["rows"][0]["subject"]["value"] == "https://example.org/item"


def test_sparql_api_queries_the_graph_saved_for_the_run(tmp_path, monkeypatch):
    run_store = RunStore(tmp_path / "runs")
    monkeypatch.setattr(playground_server, "RUN_STORE", run_store)
    state = run_store.create("items.csv", "items.csv")
    state["stages"]["ingest"] = {"status": "success"}
    state["graph"] = {
        "name": "items",
        "uri": "https://example.org/graphs/items",
    }
    run_store.save(state)
    captured = {}

    def run_query(query, graph_uri):
        captured.update({"query": query, "graph_uri": graph_uri})
        return {
            "type": "ask",
            "boolean": True,
            "graph_uri": graph_uri,
        }

    monkeypatch.setattr(generic_pipeline, "run_sparql_query", run_query)
    client = TestClient(playground_server.app)

    response = client.post(
        f"/api/runs/{state['id']}/sparql",
        json={"query": "ASK { ?subject ?predicate ?object }"},
    )

    assert response.status_code == 200
    assert response.json()["boolean"] is True
    assert captured == {
        "query": "ASK { ?subject ?predicate ?object }",
        "graph_uri": "https://example.org/graphs/items",
    }


def test_optional_validation_does_not_invalidate_downstream_results(tmp_path):
    run_store = RunStore(tmp_path / "runs")
    state = run_store.create("items.csv", "items.csv")
    state["stages"] = {
        stage: {"status": "success", "artifacts": []}
        for stage in (
            "upload",
            "rml",
            "alignment",
            "ingest",
            "shacl_in",
            "reason",
            "rdf2tss",
            "shacl_out",
            "rdf2ldes",
        )
    }
    run_store.save(state)

    state = run_store.begin_stage(state, "shacl_in")

    assert state["stages"]["shacl_in"]["status"] == "running"
    assert state["stages"]["reason"]["status"] == "success"
    assert state["stages"]["rdf2tss"]["status"] == "success"
    assert state["stages"]["rdf2ldes"]["status"] == "success"


def test_alignment_api_creates_artifact_and_ingestion_uses_it(tmp_path, monkeypatch):
    run_store = RunStore(tmp_path / "runs")
    monkeypatch.setattr(playground_server, "RUN_STORE", run_store)
    state = run_store.create("items.csv", "items.csv")
    state["stages"]["rml"] = {"status": "success"}
    run_store.save(state)
    directory = run_store.run_dir(state["id"])
    mapped_path = directory / "mapped.ttl"
    mapped_path.write_text(
        """
        @prefix qudt: <http://qudt.org/schema/qudt/> .
        @prefix sosa: <http://www.w3.org/ns/sosa/> .
        @prefix unit: <http://qudt.org/vocab/unit/> .
        <urn:observation> sosa:hasSimpleResult 1500 ; qudt:hasUnit unit:MicroS-PER-CentiM .
        """,
        encoding="utf-8",
    )
    client = TestClient(playground_server.app)

    alignment_response = client.post(
        f"/api/runs/{state['id']}/stages/alignment",
        json={"target_unit": "http://qudt.org/vocab/unit/MilliS-PER-CentiM"},
    )

    assert alignment_response.status_code == 200
    aligned_state = alignment_response.json()
    assert aligned_state["stages"]["alignment"]["status"] == "success"
    assert aligned_state["stages"]["alignment"]["details"]["converted_observations"] == 1
    assert any(artifact["name"] == "Unit-aligned RDF" for artifact in aligned_state["artifacts"])

    captured = {}

    def ingest_graph(rdf_path, graph_name):
        captured.update({"rdf_path": rdf_path, "graph_name": graph_name})
        return {"graph_uri": "https://example.org/graphs/aligned", "graph_cleared": True}

    monkeypatch.setattr(generic_pipeline, "ingest_graph", ingest_graph)
    ingest_response = client.post(
        f"/api/runs/{state['id']}/stages/ingest",
        json={"graph_name": "aligned"},
    )

    assert ingest_response.status_code == 200
    assert ingest_response.json()["stages"]["ingest"]["details"]["input_rdf"] == "aligned.ttl"
    assert captured == {"rdf_path": directory / "aligned.ttl", "graph_name": "aligned"}


def test_rdf2tss_api_uses_mapped_rdf_when_reasoning_is_skipped(tmp_path, monkeypatch):
    run_store = RunStore(tmp_path / "runs")
    monkeypatch.setattr(playground_server, "RUN_STORE", run_store)
    state = run_store.create("items.csv", "items.csv")
    state["stages"]["rml"] = {"status": "success"}
    run_store.save(state)
    directory = run_store.run_dir(state["id"])
    mapped_path = directory / "mapped.ttl"
    mapped_path.write_text("<urn:item> <urn:value> 1 .", encoding="utf-8")
    captured = {}

    def run_rdf2tss(run_directory, data_path):
        captured["data_path"] = data_path
        output_path = run_directory / "timeseries.ttl"
        output_path.write_text("<urn:series> <urn:value> 1 .", encoding="utf-8")
        return {"output_path": output_path, "sensor_count": 1, "tss_triples": 1}

    monkeypatch.setattr(generic_pipeline, "run_rdf2tss", run_rdf2tss)
    client = TestClient(playground_server.app)

    response = client.post(f"/api/runs/{state['id']}/stages/rdf2tss")

    assert response.status_code == 200
    payload = response.json()
    assert payload["stages"]["rdf2tss"]["status"] == "success"
    assert payload["stages"]["rdf2tss"]["details"]["input_rdf"] == "mapped.ttl"
    assert captured["data_path"] == mapped_path


def test_rdf2ldes_api_does_not_require_shacl_out(tmp_path, monkeypatch):
    run_store = RunStore(tmp_path / "runs")
    monkeypatch.setattr(playground_server, "RUN_STORE", run_store)
    state = run_store.create("items.csv", "items.csv")
    state["stages"]["rdf2tss"] = {"status": "success"}
    run_store.save(state)
    directory = run_store.run_dir(state["id"])
    tss_path = directory / "timeseries.ttl"
    tss_path.write_text("<urn:series> <urn:value> 1 .", encoding="utf-8")
    captured = {}

    def run_rdf2ldes(run_directory, input_path, stream_name, base_url, *, source_kind):
        captured.update({"input_path": input_path, "source_kind": source_kind})
        output_directory = run_directory / "ldes" / stream_name
        output_directory.mkdir(parents=True)
        zip_path = run_directory / f"{stream_name}.zip"
        zip_path.write_bytes(b"zip")
        return {
            "output_directory": output_directory,
            "zip_path": zip_path,
            "stream_name": stream_name,
            "base_url": base_url,
            "trig_file_count": 0,
            "fragment_count": 0,
            "index_count": 0,
            "source_file": input_path.name,
        }

    monkeypatch.setattr(generic_pipeline, "run_rdf2ldes", run_rdf2ldes)
    client = TestClient(playground_server.app)

    response = client.post(
        f"/api/runs/{state['id']}/stages/rdf2ldes",
        json={"stream_name": "items", "base_url": "https://example.org/ldes/"},
    )

    assert response.status_code == 200
    assert response.json()["stages"]["rdf2ldes"]["status"] == "success"
    assert captured["input_path"] == tss_path
    assert captured["source_kind"] == "tss"


def test_rdf2ldes_api_can_use_mapped_rdf_without_tss_stage(tmp_path, monkeypatch):
    run_store = RunStore(tmp_path / "runs")
    monkeypatch.setattr(playground_server, "RUN_STORE", run_store)
    state = run_store.create("items.csv", "items.csv")
    state["stages"]["rml"] = {"status": "success"}
    run_store.save(state)
    directory = run_store.run_dir(state["id"])
    mapped_path = directory / "mapped.ttl"
    mapped_path.write_text("<urn:item> <urn:value> 1 .", encoding="utf-8")
    captured = {}

    def run_rdf2ldes(run_directory, input_path, stream_name, base_url, *, source_kind):
        captured.update({"input_path": input_path, "source_kind": source_kind})
        output_directory = run_directory / "ldes" / stream_name
        output_directory.mkdir(parents=True)
        zip_path = run_directory / f"{stream_name}.zip"
        zip_path.write_bytes(b"zip")
        return {
            "output_directory": output_directory,
            "zip_path": zip_path,
            "stream_name": stream_name,
            "base_url": base_url,
            "trig_file_count": 1,
            "fragment_count": 1,
            "index_count": 0,
            "source_file": input_path.name,
        }

    monkeypatch.setattr(generic_pipeline, "run_rdf2ldes", run_rdf2ldes)
    client = TestClient(playground_server.app)

    response = client.post(
        f"/api/runs/{state['id']}/stages/rdf2ldes",
        json={
            "stream_name": "items",
            "base_url": "https://example.org/ldes/",
            "source": "rdf",
        },
    )

    assert response.status_code == 200
    payload = response.json()["stages"]["rdf2ldes"]
    assert payload["status"] == "success"
    assert payload["details"]["source"] == "rdf"
    assert captured == {"input_path": mapped_path, "source_kind": "rdf"}


def test_mapped_rdf_preview_paginates_by_subject(tmp_path, monkeypatch):
    run_store = RunStore(tmp_path / "runs")
    monkeypatch.setattr(playground_server, "RUN_STORE", run_store)
    state = run_store.create("items.csv", "items.csv")
    state["stages"]["rml"] = {"status": "success"}
    run_store.save(state)
    mapped_path = run_store.run_dir(state["id"]) / "mapped.ttl"
    graph = Graph()
    predicate = URIRef("https://example.org/value")
    for index in range(205):
        graph.add((URIRef(f"https://example.org/item/{index:03d}"), predicate, URIRef(f"urn:value:{index}")))
    graph.serialize(destination=mapped_path, format="turtle")
    client = TestClient(playground_server.app)

    first = client.get(f"/api/runs/{state['id']}/rdf-preview?offset=0&limit=10")
    last = client.get(f"/api/runs/{state['id']}/rdf-preview?offset=200&limit=10")

    assert first.status_code == 200
    assert first.json()["returned_instances"] == 10
    assert first.json()["total_instances"] == 205
    assert first.json()["has_previous"] is False
    assert first.json()["has_next"] is True
    assert last.json()["returned_instances"] == 5
    assert last.json()["has_previous"] is True
    assert last.json()["has_next"] is False
    assert "item/204" in last.json()["text"]


def test_shacl_violation_returns_a_report_without_raising(tmp_path):
    data_path = tmp_path / "mapped.ttl"
    data_path.write_text(
        """
        @prefix ex: <https://example.org/> .
        ex:item ex:name "Example" .
        """,
        encoding="utf-8",
    )
    shapes = """
        @prefix sh: <http://www.w3.org/ns/shacl#> .
        @prefix ex: <https://example.org/> .
        ex:ItemShape a sh:NodeShape ;
            sh:targetNode ex:item ;
            sh:property [ sh:path ex:price ; sh:minCount 1 ] .
    """

    result = generic_pipeline.run_shacl_validation(
        tmp_path, data_path, shapes, prefix="shacl_in"
    )

    assert result["conforms"] is False
    assert "Constraint Violation" in result["report"]
    assert result["report_path"].is_file()


@pytest.mark.skipif(
    not generic_pipeline.RML_MAPPER_JAR.is_file() or not shutil.which("java"),
    reason="RMLMapper and Java are required for the integration test.",
)
def test_user_supplied_rml_mapping_runs_against_uploaded_filename(tmp_path):
    source = tmp_path / "inventory.csv"
    source.write_text("id,name\n1,Desk lamp\n", encoding="utf-8")
    mapping = """
        @prefix ex: <https://example.org/> .
        @prefix ql: <http://semweb.mmlab.be/ns/ql#> .
        @prefix rml: <http://semweb.mmlab.be/ns/rml#> .
        @prefix rr: <http://www.w3.org/ns/r2rml#> .

        <#Items> a rr:TriplesMap ;
          rml:logicalSource [ rml:source "inventory.csv" ; rml:referenceFormulation ql:CSV ] ;
          rr:subjectMap [ rr:template "https://example.org/items/{id}" ; rr:class ex:Item ] ;
          rr:predicateObjectMap [ rr:predicate ex:name ; rr:objectMap [ rml:reference "name" ] ] .
    """

    result = generic_pipeline.run_rml_mapping(tmp_path, source, mapping)

    graph = Graph().parse(result["output_path"], format="turtle")
    assert (URIRef("https://example.org/items/1"), None, None) in graph


@pytest.mark.skipif(
    not generic_pipeline.RML_MAPPER_JAR.is_file() or not shutil.which("java"),
    reason="RMLMapper and Java are required for the integration test.",
)
def test_normalized_multirow_xlsx_runs_with_user_mapping(tmp_path):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Measurements"
    worksheet.append([None, None, "Ambient temperature"])
    worksheet.append(["Date", "Time", "SENSOR_A_TEMPERATURE"])
    worksheet.append(["unit", None, "degC"])
    worksheet.append([datetime(2025, 1, 1), "00:00:00", 4.679602775605543])
    source_xlsx = tmp_path / "data.xlsx"
    source_csv = tmp_path / "data.csv"
    workbook.save(source_xlsx)
    workbook.close()

    preview = generic_pipeline.xlsx_preview(source_xlsx)
    generic_pipeline.xlsx_to_csv(source_xlsx, source_csv, preview)
    mapping = """
        @base <http://example.com/observations/> .
        @prefix rr: <http://www.w3.org/ns/r2rml#> .
        @prefix rml: <http://semweb.mmlab.be/ns/rml#> .
        @prefix ql: <http://semweb.mmlab.be/ns/ql#> .
        @prefix sosa: <http://www.w3.org/ns/sosa/> .
        @prefix xsd: <http://www.w3.org/2001/XMLSchema#> .
        @prefix qudt: <http://qudt.org/schema/qudt/> .
        @prefix unit: <http://qudt.org/vocab/unit/> .
        @prefix quantitykind: <https://qudt.org/vocab/quantitykind/> .

        <#SensorMapping> a rr:TriplesMap ;
          rml:logicalSource [
            rml:source "data.csv" ;
            rml:referenceFormulation ql:CSV
          ] ;
          rr:subjectMap [
            rr:template "https://example.org/observations/temperature-1/{DateTime}" ;
            rr:class sosa:Observation
          ] ;
          rr:predicateObjectMap [
            rr:predicate sosa:madeBySensor ;
            rr:objectMap [ rr:constant <https://example.org/sensors/temperature-1> ; rr:termType rr:IRI ]
          ] ;
          rr:predicateObjectMap [
            rr:predicate sosa:resultTime ;
            rr:objectMap [ rml:reference "DateTime" ; rr:datatype xsd:dateTime ]
          ] ;
          rr:predicateObjectMap [
            rr:predicate sosa:hasSimpleResult ;
            rr:objectMap [
              rml:reference "Ambient temperature | SENSOR_A_TEMPERATURE | degC" ;
              rr:datatype xsd:double
            ]
          ] ;
          rr:predicateObjectMap [
            rr:predicate sosa:observedProperty ;
            rr:objectMap [ rr:constant quantitykind:Temperature ; rr:termType rr:IRI ]
          ] ;
          rr:predicateObjectMap [
            rr:predicate qudt:hasUnit ;
            rr:objectMap [ rr:constant unit:DEG_C ; rr:termType rr:IRI ]
          ] .
    """

    result = generic_pipeline.run_rml_mapping(tmp_path, source_csv, mapping)

    graph = Graph().parse(result["output_path"], format="turtle")
    assert result["rdf_triples"] == 6
    assert (
        URIRef("https://example.org/observations/temperature-1/2025-01-01T00%3A00%3A00"),
        URIRef("http://www.w3.org/ns/sosa/hasSimpleResult"),
        None,
    ) in graph


@pytest.mark.skipif(
    not generic_pipeline.tool_status()["eye"],
    reason="A working EYE installation is required for the integration test.",
)
def test_user_supplied_n3_rules_create_reasoned_rdf(tmp_path):
    data = tmp_path / "mapped.ttl"
    data.write_text(
        '@prefix ex: <https://example.org/> . ex:item ex:category "lamp" .',
        encoding="utf-8",
    )
    rules = """
        @prefix ex: <https://example.org/> .
        { ?item ex:category "lamp" . } => { ?item ex:illuminates true . } .
    """

    result = generic_pipeline.run_reasoner(tmp_path, data, rules)
    graph = Graph().parse(result["output_path"], format="turtle")

    assert result["inferred_triples"] >= 1
    assert (URIRef("https://example.org/item"), URIRef("https://example.org/illuminates"), None) in graph


def test_rdf2ldes_generates_a_downloadable_zip(tmp_path):
    tss_path = tmp_path / "timeseries.ttl"
    tss_path.write_text(
        '''
        @prefix rdf: <http://www.w3.org/1999/02/22-rdf-syntax-ns#> .
        @prefix sosa: <http://www.w3.org/ns/sosa/> .
        @prefix tss: <https://w3id.org/tss#> .
        @prefix xsd: <http://www.w3.org/2001/XMLSchema#> .

        <https://example.org/snippet> a tss:Snippet ;
            tss:about _:template ;
            tss:from "2026-01-02T00:00:00+00:00"^^xsd:dateTime ;
            tss:until "2026-01-02T23:59:59+00:00"^^xsd:dateTime ;
            tss:pointType sosa:Observation ;
            tss:points "[{\\"id\\": \\"https://example.org/reading/1\\", \\"time\\": \\"2026-01-02T12:00:00+00:00\\", \\"value\\": 42.0}]"^^rdf:JSON .

        _:template a tss:PointTemplate ;
            sosa:madeBySensor <https://example.org/sensor/1> ;
            sosa:observedProperty <https://example.org/property/temperature> .
        ''',
        encoding="utf-8",
    )

    result = generic_pipeline.run_rdf2ldes(
        tmp_path,
        tss_path,
        "example-stream",
        "https://data.example/ldes/",
    )

    assert result["zip_path"].is_file()
    assert result["trig_file_count"] >= 4
    assert result["fragment_count"] == 1
    assert any(path.name == "readings.trig" for path in Path(result["output_directory"]).rglob("*.trig"))


@pytest.mark.skipif(
    not generic_pipeline.RML_MAPPER_JAR.is_file() or not shutil.which("java"),
    reason="RMLMapper and Java are required for the sample integration test.",
)
def test_supplied_sample_generates_ldes_from_mapped_rdf_and_tss(tmp_path):
    repository = Path(__file__).resolve().parents[2]
    workbook_path = repository / "test-data" / "sample.xlsx"
    mapping_path = repository / "test-data" / "rml.ttl.txt"
    preview = generic_pipeline.xlsx_preview(workbook_path)
    csv_path = generic_pipeline.xlsx_to_csv(workbook_path, tmp_path / "sample.csv", preview)
    mapping = mapping_path.read_text(encoding="utf-8").replace("data.csv", "sample.csv")

    mapped = generic_pipeline.run_rml_mapping(tmp_path, csv_path, mapping)
    tss = generic_pipeline.run_rdf2tss(tmp_path, mapped["output_path"])
    from_rdf = generic_pipeline.run_rdf2ldes(
        tmp_path,
        mapped["output_path"],
        "sample-rdf",
        "https://example.org/ldes/",
        source_kind="rdf",
    )
    from_tss = generic_pipeline.run_rdf2ldes(
        tmp_path,
        tss["output_path"],
        "sample-tss",
        "https://example.org/ldes/",
        source_kind="tss",
    )

    assert mapped["rdf_triples"] > 0
    assert tss["tss_triples"] > 0
    assert from_rdf["fragment_count"] > 0
    assert from_tss["fragment_count"] > 0
    assert from_rdf["zip_path"].stat().st_size > 0
    assert from_tss["zip_path"].stat().st_size > 0
