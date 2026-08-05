"""Generic, run-scoped semantic pipeline operations.

Every function receives explicit input and output paths. Uploaded data is
materialised inside an isolated run directory, so no stage depends on the
repository's historical example data.
"""

from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import date, datetime, time as time_type
import os
from pathlib import Path
import re
import shutil
import subprocess
from threading import Lock
from typing import Any, Sequence
from urllib.parse import quote
from zipfile import BadZipFile, ZIP_DEFLATED, ZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from rdflib import Graph

from RDF2LDES import RDFTSS2LDES
from RDF2TSS_V2 import RDF2TSS_V2
from SHACL.SHACL_validate import validate_shacl
from triple_store_ingestion import ingest


ROOT_DIR = Path(__file__).resolve().parents[1]
RML_MAPPER_JAR = Path(
    os.getenv("RML_MAPPER_JAR", str(ROOT_DIR / "pipeline" / "rmlmapper.jar"))
).expanduser().resolve()
DEFAULT_GRAPH_BASE = os.getenv("FUSEKI_GRAPH_BASE", "https://example.org/graphs/")
DEFAULT_LDES_BASE = os.getenv("LDES_BASE_URL", "https://example.org/ldes/")
COMMAND_TIMEOUT_SECONDS = int(os.getenv("PIPELINE_COMMAND_TIMEOUT", "300"))
LDES_LOCK = Lock()
SUPPORTED_TABULAR_EXTENSIONS = {".csv", ".xlsx"}


class PipelineError(RuntimeError):
    """An actionable stage error safe to display in the dashboard."""

    def __init__(self, message: str, log: str = ""):
        super().__init__(message)
        self.log = log


@dataclass
class CommandResult:
    stdout: str
    stderr: str

    @property
    def log(self) -> str:
        return "\n".join(part.strip() for part in (self.stdout, self.stderr) if part.strip())


def run_command(command: Sequence[str], *, cwd: Path, label: str) -> CommandResult:
    try:
        completed = subprocess.run(
            list(command),
            cwd=cwd,
            capture_output=True,
            text=True,
            check=False,
            timeout=COMMAND_TIMEOUT_SECONDS,
        )
    except FileNotFoundError as error:
        raise PipelineError(f"{label} cannot start because an executable is missing: {error}") from error
    except subprocess.TimeoutExpired as error:
        output = "\n".join(str(item) for item in (error.stdout, error.stderr) if item)
        raise PipelineError(
            f"{label} exceeded the {COMMAND_TIMEOUT_SECONDS}-second execution limit.", output
        ) from error

    result = CommandResult(completed.stdout or "", completed.stderr or "")
    if completed.returncode != 0:
        raise PipelineError(
            f"{label} failed with exit code {completed.returncode}.",
            result.log or "The process did not provide an error message.",
        )
    return result


def safe_filename(filename: str | None) -> str:
    original = Path(filename or "dataset.csv").name
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", original).strip("._")
    if not cleaned:
        cleaned = "dataset.csv"
    if Path(cleaned).suffix.lower() not in SUPPORTED_TABULAR_EXTENSIONS:
        raise PipelineError("Please upload a .csv or .xlsx file.")
    return cleaned


def csv_preview(path: str | Path, limit: int = 30) -> dict[str, Any]:
    source = Path(path)
    sample = source.read_bytes()[:65_536]
    if not sample:
        raise PipelineError("The uploaded CSV file is empty.")

    encoding = "utf-8-sig"
    try:
        decoded_sample = sample.decode(encoding)
    except UnicodeDecodeError:
        encoding = "latin-1"
        decoded_sample = sample.decode(encoding)

    try:
        dialect = csv.Sniffer().sniff(decoded_sample, delimiters=",;\t|")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ","

    rows: list[dict[str, str]] = []
    total_rows = 0
    try:
        with source.open("r", encoding=encoding, errors="strict", newline="") as handle:
            reader = csv.DictReader(handle, delimiter=delimiter)
            columns = [str(column).strip() for column in (reader.fieldnames or []) if column is not None]
            if not columns:
                raise PipelineError("No CSV header row could be detected.")
            for raw_row in reader:
                total_rows += 1
                if len(rows) < limit:
                    rows.append({column: "" if raw_row.get(column) is None else str(raw_row[column]) for column in columns})
    except UnicodeDecodeError as error:
        raise PipelineError(f"The CSV could not be decoded as {encoding}: {error}") from error
    except csv.Error as error:
        raise PipelineError(f"The CSV could not be parsed: {error}") from error

    return {
        "format": "csv",
        "columns": columns,
        "rows": rows,
        "preview_row_count": len(rows),
        "total_rows": total_rows,
        "delimiter": "tab" if delimiter == "\t" else delimiter,
        "encoding": encoding,
        "size": source.stat().st_size,
    }


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except (TypeError, ValueError):
            pass
    return str(value)


def _unique_columns(values: Sequence[Any]) -> list[str]:
    columns: list[str] = []
    used: dict[str, int] = {}
    for index, value in enumerate(values, start=1):
        base = _cell_text(value).strip() or f"column_{index}"
        count = used.get(base, 0) + 1
        used[base] = count
        columns.append(base if count == 1 else f"{base}_{count}")
    return columns


@dataclass(frozen=True)
class XlsxColumn:
    """One output CSV column derived from one or more worksheet columns."""

    name: str
    source_indexes: tuple[int, ...]
    kind: str = "value"


@dataclass(frozen=True)
class XlsxTablePlan:
    """Detected tabular layout for an XLSX worksheet."""

    columns: tuple[XlsxColumn, ...]
    data_start_row: int
    header_row_count: int
    source_column_count: int


def _is_data_value(value: Any) -> bool:
    return value is not None and not isinstance(value, str) and isinstance(
        value, (date, time_type, int, float, bool)
    )


def _date_part(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date) and not isinstance(value, time_type):
        return value
    if isinstance(value, str):
        text = value.strip()
        try:
            return datetime.fromisoformat(text).date()
        except ValueError:
            try:
                return date.fromisoformat(text)
            except ValueError:
                return None
    return None


def _time_part(value: Any) -> time_type | None:
    if isinstance(value, datetime):
        return value.time()
    if isinstance(value, time_type):
        return value
    if isinstance(value, str):
        try:
            return time_type.fromisoformat(value.strip())
        except ValueError:
            return None
    return None


def _is_date_value(value: Any) -> bool:
    return _date_part(value) is not None


def _is_time_value(value: Any) -> bool:
    return _time_part(value) is not None


def _last_populated_index(row: Sequence[Any]) -> int:
    return max(
        (index for index, value in enumerate(row) if _cell_text(value).strip()),
        default=-1,
    )


def _xlsx_table_plan(worksheet: Any, sample_size: int = 20) -> XlsxTablePlan:
    sampled_rows: list[tuple[int, tuple[Any, ...]]] = []
    for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        if not any(_cell_text(value).strip() for value in row):
            continue
        sampled_rows.append((row_number, row))
        if len(sampled_rows) >= sample_size:
            break

    if not sampled_rows:
        raise PipelineError(f"Worksheet '{worksheet.title}' is empty.")

    first_row_number = sampled_rows[0][0]
    data_sample_index: int | None = None
    for index, (_, row) in enumerate(sampled_rows[1:], start=1):
        populated = [value for value in row if _cell_text(value).strip()]
        data_values = sum(_is_data_value(value) for value in populated)
        if populated and data_values and data_values / len(populated) >= 0.5:
            data_sample_index = index
            break

    if data_sample_index is None:
        header_rows = [sampled_rows[0][1]]
        data_start_row = first_row_number + 1
        first_data_row: tuple[Any, ...] = ()
    else:
        header_rows = [row for _, row in sampled_rows[:data_sample_index]]
        data_start_row = sampled_rows[data_sample_index][0]
        first_data_row = sampled_rows[data_sample_index][1]

    source_column_count = max(
        (_last_populated_index(row) + 1 for _, row in sampled_rows),
        default=0,
    )
    raw_columns: list[XlsxColumn] = []
    source_index = 0
    while source_index < source_column_count:
        if (
            source_index + 1 < source_column_count
            and source_index < len(first_data_row)
            and source_index + 1 < len(first_data_row)
            and _is_date_value(first_data_row[source_index])
            and _is_time_value(first_data_row[source_index + 1])
        ):
            raw_columns.append(
                XlsxColumn("DateTime", (source_index, source_index + 1), "datetime")
            )
            source_index += 2
            continue

        fragments: list[str] = []
        for header_row in header_rows:
            if source_index >= len(header_row):
                continue
            fragment = _cell_text(header_row[source_index]).strip()
            if fragment and fragment not in fragments:
                fragments.append(fragment)
        raw_columns.append(
            XlsxColumn(" | ".join(fragments) or f"column_{source_index + 1}", (source_index,))
        )
        source_index += 1

    unique_names = _unique_columns([column.name for column in raw_columns])
    columns = tuple(
        XlsxColumn(name, column.source_indexes, column.kind)
        for name, column in zip(unique_names, raw_columns, strict=True)
    )
    return XlsxTablePlan(
        columns=columns,
        data_start_row=data_start_row,
        header_row_count=len(header_rows),
        source_column_count=source_column_count,
    )


def _combined_datetime(date_value: Any, time_value: Any) -> str:
    parsed_date = _date_part(date_value)
    parsed_time = _time_part(time_value)
    if parsed_date is None or parsed_time is None:
        return ""
    return datetime.combine(parsed_date, parsed_time).isoformat()


def _xlsx_row_values(row: Sequence[Any], plan: XlsxTablePlan) -> list[str]:
    values: list[str] = []
    for column in plan.columns:
        if column.kind == "datetime":
            date_index, time_index = column.source_indexes
            date_value = row[date_index] if date_index < len(row) else None
            time_value = row[time_index] if time_index < len(row) else None
            values.append(_combined_datetime(date_value, time_value))
        else:
            source_index = column.source_indexes[0]
            values.append(_cell_text(row[source_index] if source_index < len(row) else None))
    return values


def xlsx_preview(path: str | Path, limit: int = 30) -> dict[str, Any]:
    source = Path(path)
    if not source.is_file() or source.stat().st_size == 0:
        raise PipelineError("The uploaded XLSX file is empty.")

    try:
        workbook = load_workbook(source, read_only=True, data_only=True)
    except (InvalidFileException, BadZipFile, OSError, ValueError) as error:
        raise PipelineError(f"The XLSX workbook could not be opened: {error}") from error

    try:
        worksheets = workbook.worksheets
        if not worksheets:
            raise PipelineError("The XLSX workbook does not contain a worksheet.")
        worksheet = workbook.active if workbook.active in worksheets else worksheets[0]
        plan = _xlsx_table_plan(worksheet)
        columns = [column.name for column in plan.columns]
        rows: list[dict[str, str]] = []
        total_rows = 0
        for raw_row in worksheet.iter_rows(min_row=plan.data_start_row, values_only=True):
            row_values = _xlsx_row_values(raw_row, plan)
            if not any(value.strip() for value in row_values):
                continue
            total_rows += 1
            if len(rows) < limit:
                rows.append(dict(zip(columns, row_values, strict=True)))

        return {
            "format": "xlsx",
            "columns": columns,
            "rows": rows,
            "preview_row_count": len(rows),
            "total_rows": total_rows,
            "sheet_name": worksheet.title,
            "sheet_names": workbook.sheetnames,
            "header_row_count": plan.header_row_count,
            "data_start_row": plan.data_start_row,
            "size": source.stat().st_size,
        }
    finally:
        workbook.close()


def tabular_preview(path: str | Path, limit: int = 30) -> dict[str, Any]:
    source = Path(path)
    if source.suffix.lower() == ".csv":
        return csv_preview(source, limit)
    if source.suffix.lower() == ".xlsx":
        return xlsx_preview(source, limit)
    raise PipelineError("Please upload a .csv or .xlsx file.")


def xlsx_to_csv(
    source_path: str | Path,
    destination_path: str | Path,
    preview: dict[str, Any],
) -> Path:
    source = Path(source_path)
    destination = Path(destination_path)
    try:
        workbook = load_workbook(source, read_only=True, data_only=True)
    except (InvalidFileException, BadZipFile, OSError, ValueError) as error:
        raise PipelineError(f"The XLSX workbook could not be opened: {error}") from error

    try:
        worksheet = workbook[preview["sheet_name"]]
        plan = _xlsx_table_plan(worksheet)
        destination.parent.mkdir(parents=True, exist_ok=True)
        with destination.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle)
            columns = [column.name for column in plan.columns]
            if columns != preview["columns"]:
                raise PipelineError("The XLSX worksheet structure changed while it was being parsed.")
            writer.writerow(columns)
            for raw_row in worksheet.iter_rows(min_row=plan.data_start_row, values_only=True):
                row_values = _xlsx_row_values(raw_row, plan)
                if any(value.strip() for value in row_values):
                    writer.writerow(row_values)
    finally:
        workbook.close()
    return destination


def parse_turtle(text: str, *, label: str) -> int:
    try:
        graph = Graph().parse(data=text, format="turtle", publicID="https://example.org/input/")
    except Exception as error:
        raise PipelineError(f"{label} is not valid Turtle: {error}") from error
    return len(graph)


def run_rml_mapping(
    run_directory: Path,
    source_csv: Path,
    mapping_text: str,
) -> dict[str, Any]:
    if not mapping_text.strip():
        raise PipelineError("Paste an RML mapping before running this stage.")
    mapping_document = mapping_text
    if not re.search(r"(?im)^\s*@base\s+", mapping_document):
        mapping_document = "@base <https://example.org/rml-mapping/> .\n" + mapping_document
    mapping_triples = parse_turtle(mapping_document, label="The RML mapping")
    mapping_path = run_directory / "mapping.rml.ttl"
    output_path = run_directory / "mapped.ttl"
    mapping_path.write_text(mapping_document, encoding="utf-8")

    if not RML_MAPPER_JAR.is_file():
        raise PipelineError(
            "RMLMapper is not available. Set RML_MAPPER_JAR to the mapper JAR path in the deployment."
        )

    result = run_command(
        [
            "java",
            "-jar",
            str(RML_MAPPER_JAR),
            "-m",
            str(mapping_path),
            "-o",
            str(output_path),
            "-s",
            "turtle",
        ],
        cwd=run_directory,
        label="RML mapping",
    )
    if not output_path.is_file() or output_path.stat().st_size == 0:
        raise PipelineError(
            "RMLMapper completed without producing RDF. Check that rml:source is "
            f"'{source_csv.name}' and that every template/reference name matches a CSV column.",
            result.log,
        )
    try:
        mapped_graph = Graph().parse(output_path, format="turtle")
    except Exception as error:
        raise PipelineError(f"RMLMapper produced invalid Turtle: {error}", result.log) from error
    if not mapped_graph:
        available_columns: list[str] = []
        try:
            with source_csv.open("r", encoding="utf-8-sig", newline="") as handle:
                available_columns = next(csv.reader(handle), [])
        except (OSError, csv.Error, StopIteration):
            pass
        column_hint = (
            " Available CSV columns: " + ", ".join(repr(column) for column in available_columns)
            if available_columns
            else ""
        )
        raise PipelineError(
            "RMLMapper produced zero triples. A subject template usually cannot be created when "
            "a referenced column is missing or empty." + column_hint,
            result.log,
        )
    return {
        "mapping_path": mapping_path,
        "output_path": output_path,
        "mapping_triples": mapping_triples,
        "rdf_triples": len(mapped_graph),
        "log": result.log,
    }


def graph_name_to_uri(graph_name: str) -> str:
    value = graph_name.strip()
    if not value:
        raise PipelineError("Enter a graph name before ingestion.")
    if value.startswith(("http://", "https://")):
        if any(character.isspace() for character in value):
            raise PipelineError("A graph IRI cannot contain whitespace.")
        return value
    return f"{DEFAULT_GRAPH_BASE.rstrip('/')}/{quote(value, safe='-._~')}"


def eye_command(*arguments: str) -> list[str]:
    eye_executable = os.getenv("EYE_COMMAND") or shutil.which("eye") or "eye"
    command = [eye_executable, *arguments]
    if os.name == "nt" and Path(eye_executable).suffix.lower() in {".cmd", ".bat"}:
        command = [os.environ.get("COMSPEC", "cmd.exe"), "/d", "/c", *command]
    return command


def ingest_graph(rdf_path: Path, graph_name: str) -> dict[str, Any]:
    graph_uri = graph_name_to_uri(graph_name)
    if not ingest.upload_graph(str(rdf_path), graph_uri):
        raise PipelineError(
            "Fuseki ingestion failed. Check FUSEKI_DATA_URL, the dataset name, and the Fuseki logs."
        )
    return {"graph_uri": graph_uri}


def run_shacl_validation(
    run_directory: Path,
    data_path: Path,
    shapes_text: str,
    *,
    prefix: str,
) -> dict[str, Any]:
    if not shapes_text.strip():
        raise PipelineError("Paste a SHACL shape before running validation.")
    parse_turtle(shapes_text, label="The SHACL shape")
    shapes_path = run_directory / f"{prefix}_shapes.ttl"
    report_path = run_directory / f"{prefix}_report.txt"
    shapes_path.write_text(shapes_text, encoding="utf-8")
    try:
        result = validate_shacl(str(data_path), str(shapes_path), str(report_path))
    except Exception as error:
        raise PipelineError(f"SHACL validation could not run: {error}") from error
    return {
        **result,
        "shapes_path": shapes_path,
        "report_path": report_path,
        "report": report_path.read_text(encoding="utf-8", errors="replace"),
    }


def run_reasoner(run_directory: Path, data_path: Path, rules_text: str) -> dict[str, Any]:
    if not rules_text.strip():
        raise PipelineError("Paste N3 rules before running the reasoner.")
    rules_path = run_directory / "reasoning_rules.n3"
    inferred_path = run_directory / "reasoned.ttl"
    rules_path.write_text(rules_text, encoding="utf-8")
    result = run_command(
        eye_command(str(data_path), str(rules_path), "--nope", "--pass-only-new"),
        cwd=run_directory,
        label="N3 reasoning",
    )

    try:
        combined = Graph().parse(data_path, format="turtle")
        inferred_count = 0
        if result.stdout.strip():
            inferred = Graph().parse(data=result.stdout, format="n3")
            inferred_count = len(inferred)
            combined += inferred
        combined.serialize(destination=inferred_path, format="turtle")
    except Exception as error:
        raise PipelineError(f"EYE returned RDF that could not be parsed: {error}", result.log) from error
    return {
        "rules_path": rules_path,
        "output_path": inferred_path,
        "inferred_triples": inferred_count,
        "total_triples": len(combined),
        "log": result.log,
    }


def run_rdf2tss(run_directory: Path, data_path: Path) -> dict[str, Any]:
    output_path = run_directory / "timeseries.ttl"
    try:
        source_graph = RDF2TSS_V2.load_graph(str(data_path))
        sensor_set = RDF2TSS_V2.create_sensor_set(source_graph)
        tss_graph = RDF2TSS_V2.create_tss(sensor_set, source_graph)
        RDF2TSS_V2.save_graph(str(output_path), tss_graph, overwrite=True)
    except Exception as error:
        raise PipelineError(f"RDF2TSS could not transform the RDF: {error}") from error
    if not sensor_set:
        raise PipelineError(
            "RDF2TSS found no sensors. The current query expects sosa:madeBySensor observations."
        )
    return {
        "output_path": output_path,
        "sensor_count": len(sensor_set),
        "tss_triples": len(tss_graph),
    }


def normalise_stream_name(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "-", value.strip()).strip("-._")
    if not cleaned:
        raise PipelineError("Enter an LDES stream name.")
    return cleaned


def run_rdf2ldes(
    run_directory: Path,
    tss_path: Path,
    stream_name: str,
    base_url: str | None = None,
) -> dict[str, Any]:
    safe_name = normalise_stream_name(stream_name)
    public_base = (base_url or DEFAULT_LDES_BASE).strip()
    if not public_base.startswith(("http://", "https://")):
        raise PipelineError("The LDES base URL must start with http:// or https://.")
    output_directory = run_directory / "ldes" / safe_name
    if output_directory.exists():
        shutil.rmtree(output_directory)
    output_directory.mkdir(parents=True, exist_ok=True)

    try:
        with LDES_LOCK:
            RDFTSS2LDES.generate_ldes(
                source_ttl=tss_path,
                output_directory=output_directory,
                stream_name=safe_name,
                base_url=public_base,
            )
    except Exception as error:
        raise PipelineError(f"RDF2LDES could not generate the stream: {error}") from error

    trig_files = sorted(output_directory.rglob("*.trig"))
    if not trig_files:
        raise PipelineError("RDF2LDES completed without generating any TriG files.")

    zip_path = run_directory / f"{safe_name}_ldes.zip"
    with ZipFile(zip_path, "w", compression=ZIP_DEFLATED) as archive:
        for path in trig_files:
            archive.write(path, (Path(safe_name) / path.relative_to(output_directory)).as_posix())
    return {
        "output_directory": output_directory,
        "zip_path": zip_path,
        "stream_name": safe_name,
        "base_url": public_base,
        "trig_file_count": len(trig_files),
        "fragment_count": sum(path.name == "readings.trig" for path in trig_files),
        "index_count": sum(path.name != "readings.trig" for path in trig_files),
    }


def tool_status() -> dict[str, Any]:
    eye_ready = False
    eye_detail = "EYE was not found on PATH."
    if os.getenv("EYE_COMMAND") or shutil.which("eye"):
        try:
            probe = subprocess.run(
                eye_command("--help"),
                capture_output=True,
                text=True,
                timeout=10,
            )
            eye_ready = probe.returncode == 0
            eye_detail = "" if eye_ready else (probe.stderr or probe.stdout or "EYE did not start.").strip()
        except (OSError, subprocess.TimeoutExpired) as error:
            eye_detail = str(error)
    return {
        "java": bool(shutil.which("java")),
        "eye": eye_ready,
        "eye_detail": eye_detail,
        "rml_mapper": RML_MAPPER_JAR.is_file(),
        "rml_mapper_path": str(RML_MAPPER_JAR),
    }
